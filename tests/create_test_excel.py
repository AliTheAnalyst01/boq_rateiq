import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_PATH = Path("tests/test_boq_input.xlsx")
OUTPUT_PATH.parent.mkdir(exist_ok=True)

ROWS = [
    {
        "SNO": "1",
        "DESCRIPTION": "CIVIL & INTERIOR FINISHING WORKS",
        "QTY": "",
        "UNIT": "",
        "RATE": "",
        "AMOUNT": "",
    },
    {
        "SNO": "1.1",
        "DESCRIPTION": 'Providing and laying Brick Work 4.5" Thick consisting of First Class approved bricks in cement mortar ratio 1:5 (Bestway Cement with Ravi sand) including cutting, wastage, hardware, labor, Steel Tie Bars, Scaffolding, Material Shifting, Freight & Cartage etc. Complete in all respects as per instruction of Architect.',
        "QTY": 250,
        "UNIT": "Sft",
        "RATE": "",
        "AMOUNT": "",
    },
    {
        "SNO": "1.2",
        "DESCRIPTION": "Providing and applying 12mm thick cement sand plaster on walls & columns up to any height using Bestway Cement and clean local sand, including surface preparation, wetting of masonry, smooth sponge finish, curing, scaffolding, wastage, all arrangements etc. Complete in all respects as per directions of Architect.",
        "QTY": 800,
        "UNIT": "Sft",
        "RATE": "",
        "AMOUNT": "",
    },
    {
        "SNO": "1.3",
        "DESCRIPTION": "Providing and Installing 12mm thick Gypsum Board False Ceiling water resistant (Elephant / Equivalent Brand) with aluminium T-grid and GI suspension system hung with MS rods including provision of light holes, AC holes, drop panels, access panels, complete as per design and directions of Consultant.",
        "QTY": 1200,
        "UNIT": "Sft",
        "RATE": "",
        "AMOUNT": "",
    },
    {
        "SNO": "1.4",
        "DESCRIPTION": "Providing and fixing Porcelain Floor Tile 600x600mm anti-skid, approved brand (Shabbir Tiles / Master Tiles or equivalent) including tile bond adhesive (Bostik / Mapei), approved grout, spacers, cutting, curing, cleaning, wastage, all accessories complete in all respects as per Architect directions.",
        "QTY": 500,
        "UNIT": "Sft",
        "RATE": "",
        "AMOUNT": "",
    },
    {
        "SNO": "1.5",
        "DESCRIPTION": "Providing and Fixing of Marble on already laid CC rough floor including dry bond (STILE / Approved Equivalent), approved cemented grout, spacers in required size, cutting, curing, cleaning, hacking of existing surface for bonding where necessary, wastage, all fixing accessories. (Base Rate of Marble @ 1200/- per Sft)",
        "QTY": 350,
        "UNIT": "Sft",
        "RATE": "",
        "AMOUNT": "",
    },
    {
        "SNO": "1.6",
        "DESCRIPTION": "Providing and applying two coats of acrylic emulsion paint (Dulux / Berger / Master equivalent) on plastered walls after proper wall putty and primer application including scaffolding, surface preparation, wastage complete in all respects.",
        "QTY": 1500,
        "UNIT": "Sft",
        "RATE": "",
        "AMOUNT": "",
    },
    {
        "SNO": "1.7",
        "DESCRIPTION": "Dismantling of Existing Brick Wall including breaking, removing debris from site, shifting to designated dumping area, labor, tools, equipment complete in all respects.",
        "QTY": 150,
        "UNIT": "Sft",
        "RATE": "",
        "AMOUNT": "",
    },
    {
        "SNO": "2",
        "DESCRIPTION": "HVAC WORKS",
        "QTY": "",
        "UNIT": "",
        "RATE": "",
        "AMOUNT": "",
    },
    {
        "SNO": "2.1",
        "DESCRIPTION": "Supply and Installation of Split Air Conditioner, Nominal Cooling Capacity 4.0 TR, Inverter type, approved brand (Gree / Haier / Dawlance), including copper piping with insulation, UPVC drain piping, electrical wiring from DB, testing and commissioning complete in all respects.",
        "QTY": 6,
        "UNIT": "Nos",
        "RATE": "",
        "AMOUNT": "",
    },
    {
        "SNO": "2.2",
        "DESCRIPTION": "Supply and Installation of Cassette Type Split AC 2.0 TR nominal cooling capacity, 4-way discharge, inverter compressor, approved brand including copper piping, insulation, drain piping, electrical wiring, remote control, testing commissioning complete.",
        "QTY": 4,
        "UNIT": "Nos",
        "RATE": "",
        "AMOUNT": "",
    },
    {
        "SNO": "2.3",
        "DESCRIPTION": "Providing and Fixing Copper Pipe with Armaflex insulation 13mm thick for refrigerant lines including all fittings, bends, tee connections, supports, clamps, complete in all respects.",
        "QTY": 120,
        "UNIT": "Rft",
        "RATE": "",
        "AMOUNT": "",
    },
    {
        "SNO": "3",
        "DESCRIPTION": "ELECTRICAL & ELV WORKS",
        "QTY": "",
        "UNIT": "",
        "RATE": "",
        "AMOUNT": "",
    },
    {
        "SNO": "3.1",
        "DESCRIPTION": "Supply, Installation and Commissioning of Light Circuit Wiring from MCB in DB to Switch Board, wired with 3x2.5mm sq. PVC insulated 300/500V grade cable (Pakistan Cables / Fast Cables / Newage) in 25mm heavy duty PVC conduit with all accessories, pull boxes, steel pull wires complete in all respects.",
        "QTY": 30,
        "UNIT": "Nos",
        "RATE": "",
        "AMOUNT": "",
    },
    {
        "SNO": "3.2",
        "DESCRIPTION": "Supply, Installation and Commissioning of 13A switched socket outlet (BS standard, MK / Crabtree / equivalent) with 3x4mm sq. PVC insulated cable from nearest DB, including back box recessed in wall, cover plate, complete in all respects.",
        "QTY": 45,
        "UNIT": "Nos",
        "RATE": "",
        "AMOUNT": "",
    },
    {
        "SNO": "3.3",
        "DESCRIPTION": "Supply, Installation, Testing and Commissioning of Full HD IP based POE CCTV Camera, Outdoor Day/Night type, minimum 4MP resolution, IR LED, vandal proof housing, IP-65 rated, with mounting bracket, RJ45 outlet, CAT-6 UTP cable wiring to NVR, complete in all respects.",
        "QTY": 8,
        "UNIT": "Nos",
        "RATE": "",
        "AMOUNT": "",
    },
    {
        "SNO": "3.4",
        "DESCRIPTION": "Supply and Laying of 4mm sq. PVC insulated single core copper cable (Pakistan Cables / Newage) in pre-laid conduit from Main Panel to Sub Distribution Board, including all cable accessories, ferrules, termination complete in all respects.",
        "QTY": 200,
        "UNIT": "Mtr",
        "RATE": "",
        "AMOUNT": "",
    },
    {
        "SNO": "4",
        "DESCRIPTION": "FIRE FIGHTING WORKS",
        "QTY": "",
        "UNIT": "",
        "RATE": "",
        "AMOUNT": "",
    },
    {
        "SNO": "4.1",
        "DESCRIPTION": "Supply, Installation, Testing and Commissioning of Single Loop Intelligent Addressable Fire Alarm Control Panel (FACP), expandable, with built-in power supply, NICAD battery backup, loop control module, addressable power supply, mounting chassis, complete in all respects as per specifications and drawings.",
        "QTY": 1,
        "UNIT": "Nos",
        "RATE": "",
        "AMOUNT": "",
    },
    {
        "SNO": "4.2",
        "DESCRIPTION": "Supply, Installation, Testing and Commissioning of Intelligent Addressable Optical Smoke Detector with base, address module, complete in all respects and connected to FACP.",
        "QTY": 24,
        "UNIT": "Nos",
        "RATE": "",
        "AMOUNT": "",
    },
    {
        "SNO": "5",
        "DESCRIPTION": "PLUMBING WORKS",
        "QTY": "",
        "UNIT": "",
        "RATE": "",
        "AMOUNT": "",
    },
    {
        "SNO": "5.1",
        "DESCRIPTION": "Providing and Fixing of UPVC Pipe (Popular E-Class) 32mm diameter for water supply including all fittings, elbows, tee, sockets, unions, clamps, supports, cutting, wastage, fixing labor, Material Shifting, Freight & Cartage complete in all respects.",
        "QTY": 80,
        "UNIT": "Rft",
        "RATE": "",
        "AMOUNT": "",
    },
    {
        "SNO": "5.2",
        "DESCRIPTION": "Supply and Installation of European Water Closet (WC) with soft close seat cover, concealed cistern, push plate flush, approved brand (Porta / Roca / equivalent) including all necessary fittings, brackets, silicon sealing, complete installation.",
        "QTY": 4,
        "UNIT": "Nos",
        "RATE": "",
        "AMOUNT": "",
    },
    {
        "SNO": "6",
        "DESCRIPTION": "WOODWORK & SPECIAL WORKS",
        "QTY": "",
        "UNIT": "",
        "RATE": "",
        "AMOUNT": "",
    },
    {
        "SNO": "6.1",
        "DESCRIPTION": "Providing and Installing Wall Cladding in 18mm thick MDF board with veneer finish (approved shade), mounted on MS frame 2x2 inch sections, including grooves, shadow gaps, edge profiles, complete in all respects as per Architect's design and approval.",
        "QTY": 300,
        "UNIT": "Sft",
        "RATE": "",
        "AMOUNT": "",
    },
    {
        "SNO": "6.2",
        "DESCRIPTION": "Providing and Applying Waterproofing Treatment using Pudlo / Dr. Fixit approved chemical mixed with cement plaster on all wet area walls up to 7ft height, including labor, material, surface preparation, wastage complete in all respects.",
        "QTY": 200,
        "UNIT": "Sft",
        "RATE": "",
        "AMOUNT": "",
    },
    {
        "SNO": "6.3",
        "DESCRIPTION": "Providing and Installing Aluminium Composite Panel (ACP) 4mm thick (Alucobond / Reynobond equivalent) on MS framework including fabrication, welding, anti-rust paint, fixing, joints, silicon sealant complete in all respects.",
        "QTY": 180,
        "UNIT": "Sft",
        "RATE": "",
        "AMOUNT": "",
    },
]

wb = Workbook()
ws = wb.active
ws.title = "BOQ - Test Project"

ws.column_dimensions["A"].width = 8
ws.column_dimensions["B"].width = 80
ws.column_dimensions["C"].width = 10
ws.column_dimensions["D"].width = 8
ws.column_dimensions["E"].width = 15
ws.column_dimensions["F"].width = 18

header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="1F4E79")
section_font = Font(bold=True, size=10, color="FFFFFF")
section_fill = PatternFill("solid", fgColor="2E75B6")
desc_font = Font(size=9)
sno_font = Font(bold=True, size=9)
center_align = Alignment(horizontal="center", vertical="center")
left_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

ws.merge_cells("A1:F1")
ws["A1"] = "BILL OF QUANTITIES — TEST PROJECT"
ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
ws["A1"].fill = PatternFill("solid", fgColor="1F4E79")
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30

ws.merge_cells("A2:F2")
ws["A2"] = "Interior Design & Civil Works | RateIQ Demo"
ws["A2"].font = Font(italic=True, size=10, color="595959")
ws["A2"].alignment = Alignment(horizontal="center")
ws.row_dimensions[2].height = 18

headers = ["S.No.", "Description of Item", "Qty", "Unit", "Rate (Rs.)", "Amount (Rs.)"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
ws.row_dimensions[3].height = 20

for i, row in enumerate(ROWS, start=4):
    sno = row["SNO"]
    desc = row["DESCRIPTION"]
    qty = row["QTY"]
    unit = row["UNIT"]

    try:
        is_section = float(sno) == int(float(sno))
    except:
        is_section = False

    ws.cell(row=i, column=1, value=sno).border = thin_border
    ws.cell(row=i, column=2, value=desc).border = thin_border
    ws.cell(row=i, column=3, value=qty).border = thin_border
    ws.cell(row=i, column=4, value=unit).border = thin_border
    ws.cell(row=i, column=5, value="").border = thin_border
    ws.cell(row=i, column=6, value="").border = thin_border

    if is_section:
        for col in range(1, 7):
            c = ws.cell(row=i, column=col)
            c.font = section_font
            c.fill = section_fill
            c.alignment = center_align
        ws.row_dimensions[i].height = 18
    else:
        ws.cell(row=i, column=1).font = sno_font
        ws.cell(row=i, column=1).alignment = center_align
        ws.cell(row=i, column=2).font = desc_font
        ws.cell(row=i, column=2).alignment = left_align
        ws.cell(row=i, column=3).alignment = center_align
        ws.cell(row=i, column=4).alignment = center_align
        ws.cell(row=i, column=5).alignment = center_align
        ws.cell(row=i, column=6).alignment = center_align
        ws.cell(row=i, column=5).fill = PatternFill("solid", fgColor="FFFF99")
        ws.row_dimensions[i].height = 45

last_row = len(ROWS) + 4
ws.merge_cells(f"A{last_row}:F{last_row}")
ws[f"A{last_row}"] = (
    "Note: Yellow cells (RATE column) are empty — submit this file to POST /fill-boq endpoint to get rates filled by RateIQ agent."
)
ws[f"A{last_row}"].font = Font(italic=True, size=9, color="FF0000")

wb.save(OUTPUT_PATH)
print(f"✓ Test Excel created: {OUTPUT_PATH}")
print(f"  Rows: {len(ROWS)} ({len([r for r in ROWS if r['QTY'] != ''])} line items)")
print(f"  Sections: Civil, HVAC, Electrical, Fire, Plumbing, Woodwork")
print(f"  All RATE cells: empty (yellow highlighted)")
