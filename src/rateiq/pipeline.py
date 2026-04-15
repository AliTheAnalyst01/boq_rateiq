"""
WHY:  Single entry point that wires all components together for batch processing.
      Handles the Excel I/O, parallelism, quality probes, and audit trail so
      agent.py stays focused on per-item reasoning.
WHAT: process_boq_file() — load BOQ → fill rates → save Excel with colour coding.
      initialize_all()   — create one RateIQAgent instance for the whole server.
      load_new_boq()     — read and normalise an Excel BOQ into a DataFrame.
      find_empty_rate_rows() — filter to only rows needing rate suggestions.
FITS INTO: Called by api.py (/fill-boq endpoint) and directly from CLI.
           Depends on agent.py, parser.py, market_search (for cache reset).

LIMITATION: Max 50 items in /rate-batch (enforced in api.py).
            process_boq_file() has no hardcoded limit but Tavily free tier has
            ~1000 searches/month. A 200-row BOQ = up to 600 Tavily calls.
            Monitor usage if processing large files regularly.
"""

import logging
import os
import random
import time
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from threading import Semaphore

from tenacity import retry, retry_if_exception_type, stop_after_attempt, wait_exponential

import pandas as pd
from anthropic import Anthropic
from openai import OpenAI
from tavily import TavilyClient
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side

from .agent import RateIQAgent
from .config import (
    ANTHROPIC_API_KEY,
    TAVILY_API_KEY,
    USE_OPENAI_FALLBACK,
    OPENAI_API_KEY,
    settings,
)
from .market_search import clear_market_cache
from .logging_config import (
    get_logger,
    get_operation_logger,
    set_worker_id,
    Stage,
    setup_logging,
)
from .models import BOQChunk, RateRecommendation
from .parser import find_header_row, load_all_chunks, normalize_columns
from .searcher import BOQSearcher

logger = logging.getLogger(__name__)

# ── Concurrency constants (from config — never hardcode here) ──────────────
# WHY: Semaphore prevents burst LLM calls during parallel row processing.
#      PIPELINE_SEMAPHORE_LIMIT controls the Anthropic rate-limit buffer.
_api_semaphore = Semaphore(settings.PIPELINE_SEMAPHORE_LIMIT)
_NUM_WORKERS: int = settings.PIPELINE_MAX_WORKERS

# Fraction of items to run through the LLM-as-judge quality probe
_QUALITY_PROBE_RATE = 0.20


def _decision_path(rec) -> str:
    """
    Build a compact audit string from GapAnalysis provenance fields.
    Example: "hist=sql_3word | crag=ambiguous | mkt=rejected"
    Appears in the DECISION_PATH Excel column so estimators can see *why*
    a rate is flagged without reading the full explanation.
    """
    gap = rec.gap_analysis if rec else None
    if not gap:
        return ""
    parts = [f"hist={gap.hist_source}"]
    if gap.crag_verdict not in ("unknown", "correct"):
        parts.append(f"crag={gap.crag_verdict}")
    if gap.market_rejected:
        parts.append("mkt=rejected")
    return " | ".join(parts)


def _run_quality_probe(llm, description: str, rec) -> str | None:
    """
    LLM-as-judge: runs on a random sample (~20%) of processed items.
    Returns None if PASS or not sampled; returns a warning string if anomaly found.
    Uses the light model (Haiku) — cheap enough for 20% sampling.

    Catches silent degradation that standard monitoring misses:
    e.g., Rs.0 rate, implausibly high/low values for common items.
    """
    if random.random() > _QUALITY_PROBE_RATE:
        return None
    try:
        prompt = (
            f"Pakistan construction cost QA. Is this rate suggestion plausible?\n\n"
            f"Item: {description[:80]}\n"
            f"Unit: {rec.unit}\n"
            f"Suggested Rate: Rs.{rec.suggested_rate:,.0f}\n"
            f"Confidence: {rec.confidence}\n\n"
            f"Reply with PASS or FAIL:<reason> only. No other text."
        )
        response = llm.messages.create(
            model=settings.ANTHROPIC_MODEL_LIGHT,
            max_tokens=40,
            messages=[{"role": "user", "content": prompt}],
        )
        content_blocks = list(response.content)
        text = ""
        if content_blocks and hasattr(content_blocks[0], "text"):
            text = content_blocks[0].text.strip()
        if text.upper().startswith("FAIL"):
            return text
        return None
    except Exception as exc:
        logger.debug("Quality probe error (non-fatal): %s", exc)
        return None


@retry(
    retry=retry_if_exception_type(Exception),
    stop=stop_after_attempt(2),
    wait=wait_exponential(multiplier=1, min=1, max=5),
    reraise=True,
)
def _run_agent_with_retry(agent: RateIQAgent, description: str, qty: str, unit: str) -> RateRecommendation:
    """
    WHY:  Transient Anthropic API errors (429, timeout) should retry once before
          marking the row as failed. Without this, a brief rate-limit spike causes
          the whole row to output Rs.0 with confidence=low.
    NOTE: Only 2 attempts — 3 would triple cost on systematic failures.
    """
    return agent.run(description, qty, unit)


def _process_row_worker(args: tuple) -> dict:
    """
    WHY:  Worker function for parallel row processing. Uses the pre-initialized
          agent to avoid reloading 400MB of BM25/reranker models per row.
          Semaphore limits concurrent LLM calls to respect API rate limits.
    READS: (idx, description, qty, unit, agent) tuple from executor
    WRITES: returns result dict with suggested_rate, confidence, decision_path, etc.
    """
    idx, description, qty, unit, agent = args

    try:
        set_worker_id(os.getpid() % 10000)

        with _api_semaphore:
            rec: RateRecommendation = _run_agent_with_retry(agent, description, qty, unit)
            time.sleep(1.5)  # Respect rate limits

            if rec.suggested_rate and rec.suggested_rate > 0:
                # LLM-as-judge: sampled quality probe for silent anomaly detection
                probe_result = _run_quality_probe(agent._llm, description, rec)
                if probe_result:
                    logger.warning(
                        "QA probe anomaly — '%s': %s (rate=Rs.%.0f, conf=%s)",
                        description[:40], probe_result,
                        rec.suggested_rate, rec.confidence,
                    )

                return {
                    "idx": idx,
                    "success": True,
                    "suggested_rate": rec.suggested_rate,
                    "confidence": rec.confidence,
                    "explanation": rec.explanation[:200] if rec.explanation else "",
                    "source_refs": rec.source_refs[:3],
                    "work_category": rec.work_category,
                    "search_keywords": rec.search_keywords,
                    "market_source": rec.market_rate.source_name
                    if rec.market_rate
                    else None,
                    "market_rate": rec.market_rate.contractor_rate
                    if rec.market_rate
                    else None,
                    "market_rate_url": rec.market_rate.source_url
                    if rec.market_rate
                    else None,
                    "historical_rate": max(
                        sr.chunk.rate
                        for sr in rec.search_results
                        if sr.chunk.rate and sr.chunk.rate > 0
                    )
                    if rec.search_results
                    else None,
                    "historical_source": (
                        max(
                            (sr.chunk.rate, sr.chunk.source_file)
                            for sr in rec.search_results
                            if sr.chunk.rate and sr.chunk.rate > 0
                        )[1]
                        if (
                            rec.search_results
                            and any(
                                sr.chunk.rate and sr.chunk.rate > 0
                                for sr in rec.search_results
                            )
                        )
                        else None
                    ),
                    "gap_pct": rec.gap_analysis.gap_pct if rec.gap_analysis else None,
                    "gap_verdict": rec.gap_analysis.verdict
                    if rec.gap_analysis
                    else None,
                    "decision_path": _decision_path(rec),
                }
            else:
                return {"idx": idx, "success": False, "reason": "no_rate"}

    except Exception as exc:
        return {"idx": idx, "success": False, "error": str(exc)}


def load_new_boq(filepath: Path) -> pd.DataFrame:
    """
    WHY:  Separated from process_boq_file() so validation errors surface before
          any agent calls start — expensive model loading shouldn't happen if
          the input file is broken or missing required columns.
    READS: Excel file at filepath (supports .xlsx and .xls)
    WRITES: returns normalised DataFrame; raises nothing (returns empty on failure)

    EXAMPLE:
        >>> load_new_boq(Path("data/raw/new_project_boq.xlsx"))
        DataFrame with 120 rows, RATE column mostly NaN
    """
    try:
        engine = "xlrd" if str(filepath).endswith(".xls") else "openpyxl"
        xl = pd.ExcelFile(filepath, engine=engine)

        all_dfs: list[pd.DataFrame] = []

        for sheet_name in xl.sheet_names:
            try:
                raw_df = pd.read_excel(
                    filepath, sheet_name=sheet_name, header=None, engine=engine
                )
                if raw_df.empty or len(raw_df) < 3:
                    continue

                header_row = find_header_row(raw_df)
                if header_row is None:
                    logger.debug("No header found in sheet '%s'", sheet_name)
                    continue

                df = pd.read_excel(
                    filepath, sheet_name=sheet_name, header=header_row, engine=engine
                )
                df = normalize_columns(df)
                df["_sheet_name"] = str(sheet_name)

                if "DESCRIPTION" in df.columns:
                    all_dfs.append(df)

            except Exception as sheet_exc:
                logger.warning("Sheet '%s' load failed: %s", sheet_name, sheet_exc)
                continue

        if not all_dfs:
            logger.error("No valid sheets found in %s", filepath.name)
            return pd.DataFrame()

        combined = pd.concat(all_dfs, ignore_index=True)
        logger.info("Loaded %d rows from %s", len(combined), filepath.name)
        return combined

    except Exception as exc:
        logger.error("load_new_boq failed: %s", exc)
        return pd.DataFrame()


def find_empty_rate_rows(df: pd.DataFrame) -> pd.DataFrame:
    """
    WHY:  Only rows with no RATE need agent processing. Pre-filled rows (from
          the client or a previous run) must not be overwritten.
    READS: df["RATE"] column — NaN / "" / "0" / 0 are all treated as empty.
    WRITES: returns filtered DataFrame (copy); original df is not modified.

    EXAMPLE:
        >>> find_empty_rate_rows(df_with_120_rows)
        DataFrame with 80 rows (only the unfilled ones)
    """
    try:
        if "RATE" not in df.columns:
            logger.warning("RATE column not found — returning all rows as empty.")
            return df.copy()

        mask = (
            df["RATE"].isna()
            | (df["RATE"].astype(str).str.strip() == "")
            | (df["RATE"].astype(str).str.strip() == "0")
            | (df["RATE"] == 0)
        )
        empty_df = df[mask].copy()
        logger.info(
            "find_empty_rate_rows: %d empty out of %d total rows",
            len(empty_df),
            len(df),
        )
        return empty_df

    except Exception as exc:
        logger.error("find_empty_rate_rows failed: %s", exc)
        return df.copy()


def _apply_excel_formatting(output_path: Path) -> None:
    """
    WHY:  An estimator scanning 120 rows needs visual triage — red rows demand
          attention before submission, green rows can be trusted immediately.
          Colour-coding CONFIDENCE and GAP_VERDICT provides that at a glance.
    READS: output_path (already written Excel file)
    WRITES: modifies the file in-place with fills, borders, column widths.
    NOTE: Formatting is cosmetic — a formatting failure must never crash the pipeline.
          The caller wraps this in try/except.
    """
    try:
        wb = load_workbook(output_path)
        ws = wb.active

        confidence_colors = {
            "high": "C6EFCE",
            "medium": "FFEB9C",
            "low": "FFCC99",
        }
        verdict_colors = {
            "CONFIRMED": "C6EFCE",
            "MINOR_GAP": "FFEB9C",
            "MAJOR_GAP": "FFC7CE",
            "NO_MARKET_DATA": "D9D9D9",
        }

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            confidence_cell = None
            verdict_cell = None

            for cell in row:
                if cell.column == 7:  # CONFIDENCE column (G)
                    confidence_cell = cell
                elif cell.column == 11:  # GAP_VERDICT column (K)
                    verdict_cell = cell

                cell.border = thin_border
                if cell.alignment and not cell.alignment.wrap_text:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

            if confidence_cell and confidence_cell.value:
                conf_val = str(confidence_cell.value).lower().strip()
                if conf_val in confidence_colors:
                    confidence_cell.fill = PatternFill(
                        "solid", fgColor=confidence_colors[conf_val]
                    )

            if verdict_cell and verdict_cell.value:
                verd_val = str(verdict_cell.value).strip()
                if verd_val in verdict_colors:
                    verdict_cell.fill = PatternFill(
                        "solid", fgColor=verdict_colors[verd_val]
                    )

        col_widths = {
            "A": 8,
            "B": 80,
            "C": 10,
            "D": 8,
            "E": 12,
            "F": 15,
            "G": 12,
            "H": 50,
            "I": 35,
            "J": 15,
            "K": 15,
            "L": 12,
            "M": 18,
            "N": 35,   # DECISION_PATH
        }
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        wb.save(output_path)
        logger.info("Applied formatting to %s", output_path)

    except Exception as e:
        logger.warning("Formatting failed: %s", e)


def process_boq_file(
    input_path: Path,
    output_path: Path,
    agent: RateIQAgent,
) -> dict:
    """
    WHY:  Top-level orchestrator for batch BOQ processing. Coordinates loading,
          parallel agent calls, result writing, confidence audit, and Excel formatting.
    READS: input_path (Excel BOQ), agent (pre-initialized RateIQAgent)
    WRITES: output_path (Excel with SUGGESTED_RATE, CONFIDENCE, DECISION_PATH, etc.)
    Returns: {"total_rows", "filled", "skipped", "output_file"}

    EXAMPLE:
        >>> process_boq_file(Path("new_boq.xlsx"), Path("filled_boq.xlsx"), agent)
        {"total_rows": 120, "filled": 95, "skipped": 25, "output_file": "..."}
    """
    result = {
        "total_rows": 0,
        "filled": 0,
        "skipped": 0,
        "output_file": str(output_path),
    }

    try:
        # Clear market cache at start of each batch so stale rates don't carry over
        clear_market_cache()

        df = load_new_boq(input_path)
        if df.empty:
            logger.error("No data loaded from %s", input_path)
            return result

        result["total_rows"] = len(df)

        empty_rows = find_empty_rate_rows(df)
        logger.info("Processing %d empty rate rows...", len(empty_rows))

        df["SUGGESTED_RATE"] = None
        df["CONFIDENCE"] = None
        df["EXPLANATION"] = None
        df["SOURCES"] = None
        df["WORK_CATEGORY"] = None
        df["SEARCH_KEYWORDS"] = None
        df["MARKET_SOURCE"] = None
        df["MARKET_RATE"] = None
        df["MARKET_RATE_URL"] = None
        df["HISTORICAL_RATE"] = None
        df["HISTORICAL_SOURCE"] = None
        df["GAP_PCT"] = None
        df["GAP_VERDICT"] = None
        df["DECISION_PATH"] = None

        work_items = []
        for idx in empty_rows.index:
            row = df.loc[idx]
            description = str(row.get("DESCRIPTION", "")).strip()
            qty = str(row.get("QTY", "")).strip()
            unit = str(row.get("UNIT", "")).strip()

            if description and len(description) >= 5:
                work_items.append((idx, description, qty, unit))

        logger.info(
            "Processing %d rows with %d workers...",
            len(work_items),
            _NUM_WORKERS,
        )

        op_logger = get_operation_logger()
        op_logger.start(len(work_items), "BOQ Rate Filling")

        results_list = []
        completed_count = 0

        with ThreadPoolExecutor(max_workers=_NUM_WORKERS) as executor:
            futures = {
                executor.submit(
                    _process_row_worker, (idx, description, qty, unit, agent)
                ): (idx, description)
                for idx, description, qty, unit in work_items
            }
            for future in as_completed(futures):
                try:
                    res = future.result()
                    results_list.append(res)
                    completed_count += 1

                    is_error = not res.get("success")
                    op_logger.update(
                        completed=completed_count,
                        stage=Stage.COMPLETE if not is_error else Stage.ERROR,
                        current_item=res.get("suggested_rate", "N/A"),
                        failed=is_error,
                    )

                except Exception as exc:
                    idx, desc = futures[future]
                    logger.error("Row %d failed: %s", idx, exc)
                    results_list.append(
                        {"idx": idx, "success": False, "error": str(exc)}
                    )
                    completed_count += 1
                    op_logger.update(
                        completed=completed_count, stage=Stage.ERROR, failed=True
                    )

        op_logger.set_stage(Stage.SAVING)
        logger.info("Writing results to Excel...")

        for res in results_list:
            idx = res["idx"]
            if res.get("success"):
                df.at[idx, "SUGGESTED_RATE"] = res.get("suggested_rate")
                df.at[idx, "CONFIDENCE"] = res.get("confidence")
                df.at[idx, "EXPLANATION"] = res.get("explanation")
                df.at[idx, "SOURCES"] = ", ".join(res.get("source_refs", []))
                df.at[idx, "WORK_CATEGORY"] = res.get("work_category")
                df.at[idx, "SEARCH_KEYWORDS"] = res.get("search_keywords")
                df.at[idx, "MARKET_SOURCE"] = res.get("market_source")
                df.at[idx, "MARKET_RATE"] = res.get("market_rate")
                df.at[idx, "MARKET_RATE_URL"] = res.get("market_rate_url")
                df.at[idx, "HISTORICAL_RATE"] = res.get("historical_rate")
                df.at[idx, "HISTORICAL_SOURCE"] = res.get("historical_source")
                df.at[idx, "GAP_PCT"] = res.get("gap_pct")
                df.at[idx, "GAP_VERDICT"] = res.get("gap_verdict")
                df.at[idx, "DECISION_PATH"] = res.get("decision_path", "")
                result["filled"] += 1
            else:
                result["skipped"] += 1
                reason = res.get("reason") or res.get("error", "unknown")
                logger.warning("  [%d] %s", idx, reason)

        # Batch confidence summary — surfaces quality signal before saving
        filled_results = [r for r in results_list if r.get("success")]
        if filled_results:
            conf_counts = Counter(r.get("confidence", "low") for r in filled_results)
            total_filled = len(filled_results)
            logger.info(
                "Batch confidence: high=%d (%.0f%%) | medium=%d (%.0f%%) | low=%d (%.0f%%)",
                conf_counts["high"],   conf_counts["high"]   / total_filled * 100,
                conf_counts["medium"], conf_counts["medium"] / total_filled * 100,
                conf_counts["low"],    conf_counts["low"]    / total_filled * 100,
            )
            if conf_counts["low"] / total_filled > 0.30:
                logger.warning(
                    "%.0f%% of rows have low confidence — "
                    "review DECISION_PATH column before submitting BOQ",
                    conf_counts["low"] / total_filled * 100,
                )

        output_path.parent.mkdir(parents=True, exist_ok=True)
        df.to_excel(output_path, index=False, engine="openpyxl")
        _apply_excel_formatting(output_path)

        logger.info(
            "Saved output: %s | filled=%d, skipped=%d",
            output_path,
            result["filled"],
            result["skipped"],
        )

        stats = op_logger.finish(success=True)
        op_logger.log_summary(stats)

        return result

    except Exception as exc:
        logger.error("process_boq_file failed: %s", exc)
        result["output_file"] = ""
        return result


def _create_llm_client() -> Anthropic | OpenAI:
    """
    WHY:  Centralised LLM client creation with OpenAI fallback. If Anthropic key
          is exhausted or invalid, the pipeline degrades to GPT-4o rather than
          failing completely. This is the only place where this logic lives.
    """
    if ANTHROPIC_API_KEY:
        try:
            client = Anthropic(api_key=ANTHROPIC_API_KEY)
            logger.info("Using Anthropic LLM")
            return client
        except Exception as anthropic_err:
            logger.warning(f"Anthropic failed ({anthropic_err}), trying OpenAI...")

    if OPENAI_API_KEY:
        try:
            openai_client = OpenAI(api_key=OPENAI_API_KEY)
            openai_client.models.list()
            logger.info("Using OpenAI GPT-4o")
            return _OpenAIFallbackClient(openai_client)
        except Exception as e:
            raise RuntimeError(
                f"Both Anthropic and OpenAI failed. Anthropic: {anthropic_err}, OpenAI: {e}"
            )

    raise RuntimeError(
        "No LLM API key configured. Set OPENAI_API_KEY or ANTHROPIC_API_KEY"
    )


class _OpenAIFallbackClient:
    """Wrapper to make OpenAI client compatible with Anthropic interface."""

    def __init__(self, openai_client, model="gpt-4o"):
        self._client = openai_client
        self._model = model

    def messages(self):
        return _OpenAIMessages(self._client, self._model)


class _OpenAIMessages:
    def __init__(self, client, model):
        self._client = client
        self._model = model

    def create(self, model, max_tokens, messages):
        response = self._client.chat.completions.create(
            model=self._model, max_tokens=max_tokens, messages=messages
        )
        return _OpenAIResponse(response)


class _OpenAIResponse:
    def __init__(self, response):
        self._response = response

    @property
    def content(self):
        return self._response.choices[0].message.content


def initialize_all(chunks_csv: Path = settings.CHUNKS_CSV) -> RateIQAgent:
    """
    INPUT:  Path to boq_chunks.csv (defaults to settings value)
    OUTPUT: RateIQAgent — fully initialized with all dependencies

    Initializes in order:
    1. Load chunks from CSV → list[BOQChunk]
    2. BOQSearcher(chunks) → builds BM25 + loads reranker
    3. Anthropic(api_key=ANTHROPIC_API_KEY)
    4. TavilyClient(api_key=TAVILY_API_KEY)
    5. RateIQAgent(searcher, llm, tavily)

    This is the ONLY function called from outside.
    Call once at app startup.

    EXAMPLE:
        >>> input:  Path("data/processed/boq_chunks.csv")
        >>> output: RateIQAgent ready to call .run() on BOQ rows
    """
    try:
        logger.info("Step 1/5: Loading chunks from %s ...", chunks_csv)
        chunks: list[BOQChunk] = load_all_chunks(chunks_csv)
        if not chunks:
            raise RuntimeError(f"No chunks loaded from {chunks_csv}")
        logger.info("  Loaded %d chunks.", len(chunks))

        logger.info("Step 2/5: Initializing BOQSearcher (BM25 + reranker)...")
        searcher = BOQSearcher(chunks)

        logger.info("Step 3/5: Initializing LLM (Anthropic with OpenAI fallback)...")
        llm = _create_llm_client()

        logger.info("Step 4/5: Initializing Tavily client...")
        tavily = TavilyClient(api_key=TAVILY_API_KEY)

        logger.info("Step 5/5: Building RateIQAgent...")
        agent = RateIQAgent(searcher=searcher, llm=llm, tavily=tavily)

        logger.info("initialize_all complete. Agent ready.")
        return agent

    except Exception as exc:
        logger.error("initialize_all failed: %s", exc)
        raise


if __name__ == "__main__":
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)-8s | %(name)s | %(message)s",
    )

    agent = initialize_all()

    rec = agent.run(
        description='Providing and laying Brick Work 4.5" Thick in CM 1:4',
        qty="1000",
        unit="sft",
    )

    logger.info("=" * 60)
    logger.info("Smoke test result:")
    logger.info("  Description : %s", rec.description[:60])
    logger.info("  Rate        : Rs. %.0f / %s", rec.suggested_rate, rec.unit)
    logger.info("  Confidence  : %s", rec.confidence)
    logger.info("  Category    : %s", rec.work_category)
    logger.info("  Keywords    : %s", rec.search_keywords)
    logger.info("  Sources     : %s", rec.source_refs)
    logger.info("  Explanation : %s", rec.explanation)
    logger.info("=" * 60)
